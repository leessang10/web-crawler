from openpyxl import load_workbook

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("C:/Users/dawin07/Documents/Cretop_Data_Crawler/data.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['최종본']

# 지정한 셀의 값 출력
cnt = 0
table = load_ws["E4":"AD4"]
for row in get_cells:
    for cell in row:
        if cnt == 0:
            print("사업자등록번호")
        if cnt == 1:
            print("총자산")
        print(cell.value, end=" ")
        cnt += 1
    print()

# # 모든 행 단위로 출력
#
# for row in load_ws.rows:
#     print(row)
#
# # 모든 열 단위로 출력
#
# for column in load_ws.columns:
#     print(column)
#
# # 모든 행과 열 출력
#
# all_values = []
# for row in load_ws.rows:
#     row_value = []
#     for cell in row:
#         row_value.append(cell.value)
#     all_values.append(row_value)
# print(all_values)
#
# load_ws.cell(3, 3, 51470)
# load_ws.cell(4, 3, 21470)
# load_ws.cell(5, 3, 1470)
# load_ws.cell(6, 3, 6470)
load_wb.save('C:/Users/dawin07/Documents/Cretop_Data_Crawler/output.xlsx')
